from __future__ import annotations
import itertools
from dataclasses import dataclass
from typing import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from openpyxl.worksheet.worksheet import Worksheet

from app.models import TotalStatistic, UserBetsStat
from app.models.config.currency import CurrenciesConfig
from app.models.statistic.full_user_stats import FullUserStat
from app.models.statistic.thread_users import ThreadUsers
from app.models.statistic.transaction import TransactionStatData, TransactionStatCaptions
from app.models.statistic.user_stats import UserStatCaptions
from app.services.collections_utils import get_first_dict_value
from app.services.reports.common import excel_bets_caption_name


@dataclass
class CellAddress:
    row: int
    column: int

    @property
    def kwargs(self, ):
        return dict(row=self.row, column=self.column)

    def shift(self, *, row: int = 0, column: int = 0) -> CellAddress:
        return CellAddress(row=self.row + row, column=self.column + column)

    def replace(self, *, row: int = None, column: int = None):
        return CellAddress(
            row=self.row if row is None else row,
            column=self.column if column is None else column
        )


A1 = CellAddress(1, 1)
TRANSACTION_FIRST_ROW = 1
"""from 0"""
TRANSACTION_FIRST_COL_SHIFT = 1


class ExcelWriter:
    number_format = r"#,##0.00\ [$CURRENCY];[Red]-#,##0.00\ [$CURRENCY]"
    date_columns = (1, )
    name_columns = (3, )

    def __init__(self, config: CurrenciesConfig):
        self.wb = Workbook()
        self._remove_all_worksheets()
        self.current_currency = config.default_currency

        self.user_currencies_columns = {self.current_currency.symbol: (7, 8, 9)}
        self.user_local_currencies_columns = (4, 5, 6)
        self.user_bookmaker_name_col = 10
        self.user_names_cols = [self.user_bookmaker_name_col]

    def insert_total_report(self, report_data: dict[int, TotalStatistic]):
        total_ws = self.wb.create_sheet("Общая сводка матчей")
        _insert_row(total_ws, get_first_dict_value(report_data).get_captions(), A1)
        currencies_columns = {self.current_currency.symbol: (4, 5, 6)}

        for i, report_row in enumerate(sorted(report_data.values(), key=lambda ts: ts.thread.id), 1):
            _insert_row(total_ws, report_row.get_printable(), A1.shift(row=i))
            self.format_rows(total_ws, A1.shift(row=i), self.date_columns, currencies_columns)
        _make_auto_width(
            total_ws,
            len(get_first_dict_value(report_data).get_printable()),
            self.date_columns,
            itertools.chain.from_iterable(currencies_columns.values()),
            self.name_columns,
        )

    def insert_thread_users(self, report_data: list[ThreadUsers]):
        thread_users_ws = self.wb.create_sheet("Отчёты от работников")
        _insert_row(thread_users_ws, report_data[0].get_captions(), A1)
        for i, report_row in enumerate(report_data, 1):
            _insert_row(thread_users_ws, report_row.get_printable(), A1.shift(row=i))
            self.format_rows(thread_users_ws, A1.shift(row=i), self.date_columns, {})
        _make_auto_width(thread_users_ws, len(report_data[0].get_printable()), self.date_columns, {}, self.name_columns)

    def insert_users_reports(self, report_data: dict[int, FullUserStat], config: CurrenciesConfig):
        # sorted by user_id for consistent sheets order
        # after sorting user_ids (dict key) don't need anymore
        sorted_reports: Iterable[FullUserStat] = map(lambda x: x[1], sorted(report_data.items(), key=lambda x: x[0]))

        for report_by_user in sorted_reports:
            user = report_by_user.get_user()
            if user is None:
                continue
            sheet = self.wb.create_sheet(excel_bets_caption_name(user))
            self.write_user_bets_report(sheet, report_by_user.bets)
            self.write_user_transaction(sheet, report_by_user.transactions, config)

    def write_user_bets_report(self, sheet: Worksheet, report_by_user: list[UserBetsStat]):
        column_count = len(UserStatCaptions.get_captions())
        _make_auto_width(
            sheet,
            column_count,
            self.date_columns,
            currencies=itertools.chain.from_iterable(self.user_currencies_columns.values()),
            names=[*self.user_names_cols, *self.name_columns],
        )
        _insert_row(sheet, UserStatCaptions.get_captions(), A1)
        if not report_by_user:
            # even if report is empty - must render header
            return
        for i, report_row in enumerate(report_by_user, 1):
            _insert_row(sheet, report_row.get_printable(), A1.shift(row=i))
            self.format_rows(
                sheet,
                A1.shift(row=i),
                self.date_columns,
                currencies=update_currency_dictionary(
                    self.user_currencies_columns, {report_row.currency.symbol: self.user_local_currencies_columns})
            )

    def write_user_transaction(
        self, sheet: Worksheet, transactions: list[TransactionStatData], config: CurrenciesConfig,
    ):
        columns = TransactionStatCaptions(first_col=len(UserStatCaptions.get_captions()) + 3)
        first_cell = A1.shift(column=columns.offset, row=TRANSACTION_FIRST_ROW)
        _make_auto_width(
            sheet=sheet,
            count=columns.get_count_columns(),
            date_columns=columns.get_date_columns(),
            currencies=columns.get_all_currency_columns(),
            names=columns.get_names_columns(),
            offset=columns.offset,
        )
        _insert_row(sheet, columns.get_captions(), first_cell)
        if not transactions:
            # even if transactions is empty - must render header
            return
        for i, transaction in enumerate(transactions, 1):
            _insert_row(sheet, transaction.get_printable(), first_cell.shift(row=i))
            self.format_rows(
                sheet=sheet,
                first_cell=first_cell.shift(row=i),
                date_columns=columns.get_date_columns(),
                currencies=columns.get_currencies_columns(
                    transaction.currency.symbol, self.current_currency.symbol,
                ),
            )
        sums = {
            cur: sum(
                map(
                    lambda t: t.amount,
                    filter(lambda t: t.currency.iso_code, transactions)
                )
            ) for cur in sorted({t.currency.iso_code for t in transactions})
        }
        sum_eur = sum(map(lambda t: t.amount_eur, transactions))

        first_balance_cell = first_cell.shift(row=-1, column=1)
        _insert_row(sheet, [f"Баланс в {config.default_currency.iso_code}", sum_eur], first_balance_cell)
        self.format_rows(
            sheet,
            first_balance_cell,
            [],
            {config.default_currency.symbol: [columns.patch_column_by_offset(2 + TRANSACTION_FIRST_COL_SHIFT)]},
        )

    def save(self, destination):
        self.wb.save(destination)

    def _remove_all_worksheets(self):
        for ws in self.wb.worksheets:
            self.wb.remove(ws)

    def format_rows(
            self,
            sheet: Worksheet,
            first_cell: CellAddress,
            date_columns: Iterable[int],
            currencies: dict[str, Iterable[int]],
    ):
        for i in date_columns:
            cell = sheet.cell(**first_cell.replace(column=i).kwargs)
            cell.number_format = numbers.FORMAT_DATE_DDMMYY  # noqa
        for currency, columns in currencies.items():
            for i in columns:
                cell = sheet.cell(**first_cell.replace(column=i).kwargs)
                cell.number_format = self.number_format.replace("CURRENCY", currency)  # noqa


def _make_auto_width(
        sheet: Worksheet,
        count: int,
        date_columns: Iterable[int],
        currencies: Iterable[int],
        names: Iterable[int] = tuple(),
        offset: int = 0,
):
    for i in range(1 + offset, offset + count + 2):
        sheet.column_dimensions[get_column_letter(i)].auto_size = True
    for i in date_columns:
        sheet.column_dimensions[get_column_letter(i)].width += -3
    for i in names:
        sheet.column_dimensions[get_column_letter(i)].width += 15
    if "Общая сводка матчей" in sheet.title:
        width_add = 15
    else:
        width_add = 3
    for i in currencies:
        sheet.column_dimensions[get_column_letter(i)].width += width_add


def _insert_row(sheet: Worksheet, data: list[str], first_cell: CellAddress):
    for i, text in enumerate(data):
        sheet.cell(**first_cell.shift(column=i).kwargs, value=text)


def update_currency_dictionary(d1: dict[str, Iterable[int]], d2: dict[str, Iterable[int]]):
    result = dict(**d1)
    for key, value in d2.items():
        if key in result:
            result[key] = [*result[key], *d2[key]]
        else:
            result[key] = value
    return result
